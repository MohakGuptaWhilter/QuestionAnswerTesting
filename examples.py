"""
Example usage of the PDF Processor and Quiz Evaluator.
This script demonstrates how to use both main functions of the QA Testing System.
"""

from src.pdf_processor import PDFProcessor
from src.quiz_evaluator import QuizEvaluator
import os
import json


def setup_example_data():
    """Create example Excel file with sample questions for testing."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    print("\n📝 Setting up example data...")
    
    # Create output directory
    os.makedirs("data/output", exist_ok=True)
    
    # Create example Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Q&A"
    
    # Add headers
    headers = ["Question #", "Question", "Correct Answer"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add sample questions
    sample_questions = [
        (1, "What is the capital of France?", "A"),
        (2, "Which planet is closest to the Sun?", "B"),
        (3, "What is the chemical symbol for Gold?", "D"),
        (4, "Who wrote Romeo and Juliet?", "C"),
        (5, "What is the largest ocean on Earth?", "A"),
    ]
    
    for q_id, question, answer in sample_questions:
        ws.append([q_id, question, answer])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18
    
    # Center align
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[2].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Save
    output_file = "data/output/example_questions.xlsx"
    wb.save(output_file)
    print(f"✓ Created example Excel file: {output_file}")
    
    return output_file


def example_1_pdf_processor():
    """
    Example 1: PDF Processing
    Extract questions and answers from PDFs and export to Excel.
    """
    print("\n" + "="*70)
    print("EXAMPLE 1: PDF Processor - Extract Questions & Answers from PDFs")
    print("="*70)
    
    print("\n📌 Usage:")
    print("""
    from src.pdf_processor import PDFProcessor
    
    # Step 1: Initialize processor with PDF paths
    processor = PDFProcessor(
        questions_pdf_path="data/input/questions.pdf",
        answers_pdf_path="data/input/answers.pdf"
    )
    
    # Step 2: Process PDFs and export to Excel
    output_file = processor.process_and_export("data/output/extracted_qa.xlsx")
    
    # Step 3: Get summary statistics
    summary = processor.get_summary()
    print(f"Questions extracted: {summary['total_questions']}")
    print(f"Answers extracted: {summary['total_answers']}")
    """)
    
    print("\n📊 Expected Output:")
    print("  - Excel file with columns: Question #, Question, Correct Answer")
    print("  - Formatted with headers and optimal column widths")
    print("\n💡 How it works:")
    print("  1. Reads PDF files using PyPDF2")
    print("  2. Parses questions (numbered format or Q: prefix)")
    print("  3. Parses answers (Answer: X format or A/B/C/D)")
    print("  4. Creates Excel workbook with formatted output")


def example_2_quiz_evaluator():
    """
    Example 2: Quiz Evaluator
    Test answers via API and generate results report.
    """
    print("\n" + "="*70)
    print("EXAMPLE 2: Quiz Evaluator - Test Answers & Generate Report")
    print("="*70)
    
    print("\n📌 Usage:")
    print("""
    from src.quiz_evaluator import QuizEvaluator
    
    # Step 1: Initialize evaluator with Excel file and API endpoint
    evaluator = QuizEvaluator(
        excel_file_path="data/output/example_questions.xlsx",
        api_endpoint="http://your-api.com/api/validate"
    )
    
    # Step 2: Load questions from Excel
    questions = evaluator.load_questions_from_excel()
    
    # Step 3: Provide user answers (in same order as Excel)
    user_answers = ["A", "B", "C", "D", "A"]
    
    # Step 4: Test answers via API
    results = evaluator.test_answers(user_answers)
    
    # Step 5: Export results to Excel with color coding
    output_file = evaluator.export_results_to_excel("data/output/test_results.xlsx")
    
    # Step 6: Get statistics
    summary = evaluator.get_summary()
    print(f"Accuracy: {summary['accuracy']:.1f}%")
    """)
    
    print("\n📊 Expected Output:")
    print("  - Excel file with original questions + test results")
    print("  - Columns: Question #, Question, Correct Answer, User Answer, Result, Expected Answer, Error")
    print("  - Color-coded results: Green=Correct, Red=Incorrect, Yellow=Error")
    print("\n💡 How it works:")
    print("  1. Loads questions from Excel file")
    print("  2. Makes POST requests to API endpoint with each answer")
    print("  3. Collects API responses (correct/incorrect)")
    print("  4. Generates Excel report with color-coded results")
    print("  5. Calculates accuracy statistics")


def run_example_workflow():
    """
    Run a complete workflow example using the QuizEvaluator.
    """
    print("\n" + "="*70)
    print("EXAMPLE 3: Complete Workflow - Test with Mock API")
    print("="*70)
    
    try:
        # Setup example data
        excel_file = setup_example_data()
        
        # Load the evaluator (note: API endpoint won't work without actual server)
        print("\n🔧 Loading evaluator with example data...")
        evaluator = QuizEvaluator(
            excel_file_path=excel_file,
            api_endpoint="http://localhost:5000/api/validate"
        )
        
        # Load questions
        questions = evaluator.load_questions_from_excel()
        print(f"✓ Loaded {len(questions)} questions")
        
        # Show sample questions
        print("\n📋 Sample Questions:")
        for q in questions[:3]:
            print(f"  Q{q['id']}: {q['question']}")
            print(f"     Answer: {q['correct_answer']}\n")
        
        print("⚠️  To test answers via API:")
        print("   1. Start your API server at http://localhost:5000")
        print("   2. Provide user answers as a list")
        print("   3. Call evaluator.test_answers(user_answers)")
        print("   4. Call evaluator.export_results_to_excel()")
        
    except Exception as e:
        print(f"✗ Error in example: {e}")


def main():
    """Main function to display all examples."""
    print("\n" + "="*70)
    print("QA TESTING SYSTEM - USAGE EXAMPLES")
    print("="*70)
    
    # Show PDF processor example
    example_1_pdf_processor()
    
    # Show quiz evaluator example
    example_2_quiz_evaluator()
    
    # Run complete workflow
    run_example_workflow()
    
    # Summary
    print("\n" + "="*70)
    print("QUICK REFERENCE")
    print("="*70)
    
    print("""
Function 1: PDFProcessor class
├─ process_and_export(output_path) → Exports to Excel
├─ extract_text_from_pdf(pdf_path) → Returns extracted text
├─ parse_questions(text) → Returns list of questions
├─ parse_answers(text) → Returns list of answers
└─ get_summary() → Returns statistics

Function 2: QuizEvaluator class
├─ load_questions_from_excel() → Returns questions from Excel
├─ test_answers(user_answers) → Tests via API, returns results
├─ export_results_to_excel(output_path) → Exports results
├─ validate_answer_with_api(q_id, answer) → Tests single answer
└─ get_summary() → Returns accuracy statistics

Required Files/Setup:
✓ requirements.txt - Dependencies (PyPDF2, openpyxl, requests)
✓ .env (optional) - Environment variables for API keys
✓ API endpoint - Must accept POST with {question_id, user_answer}

Directory Structure:
data/
├─ input/     ← Place PDF files here
└─ output/    ← Excel files saved here

For detailed usage, see README.md
    """)


if __name__ == "__main__":
    main()
