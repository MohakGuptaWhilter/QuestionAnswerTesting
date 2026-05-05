import os
import re
import requests as http_requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


FIGURE_URL_RE = re.compile(r'\[\[FIGURE_URL:([^\]]+)\]\]')
FIG_FONT = Font(color='0070C0', underline='single')


def inline_fig_labels(question: str) -> str:
    """Replace each [[FIGURE_URL:...]] with [FIGURE1], [FIGURE2], … in reading order."""
    counter = 0
    def _sub(_):
        nonlocal counter
        counter += 1
        return f'[FIGURE{counter}]'
    return FIGURE_URL_RE.sub(_sub, question)


SEARCH_API_URL = "https://dev.api.kb.whilter.ai/api/search"
_AUTH_TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiMGZkY2M5YmMtYzY0OC00YzRiLWE5NDUtNWE1NGJhZmYyZWI1IiwiZXhwIjoxNzc4OTkyNjI4fQ.vXssK0Dg3xWdOoVnq3zlZ1txrXtzjy7vZlQneeUmqHU"

_ILLEGAL_EXCEL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f￾￿]')

NOISE_PATTERNS = [
    re.compile(r'<a\s[^>]*>.*?</a>', re.IGNORECASE | re.DOTALL),
    re.compile(r'<!--.*?-->', re.DOTALL),
    re.compile(r'Text\s*&\s*Video Solutions.*', re.IGNORECASE),
    re.compile(r'Download MARKS App.*', re.IGNORECASE),
    re.compile(r'https?://\S+'),
    re.compile(r'Mathematics Top \d+ PYQs.*', re.IGNORECASE),
    re.compile(r'^MathonGo\s*$', re.IGNORECASE | re.MULTILINE),
]


def sanitize(text: str) -> str:
    return _ILLEGAL_EXCEL_CHARS.sub('', text) if text else text


def clean_question(text: str) -> str:
    for pattern in NOISE_PATTERNS:
        text = pattern.sub('', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def call_search_api(question: str, agent_id: str, deployment_slug: str) -> str:
    payload = {
        "query": question,
        "agent_id": agent_id,
        "tenant_id": "e6475d6e-f357-443f-8ab9-f0f61081191e",
        "top_k": 5,
        "filters": {"additionalProp1": {}},
        "chat_history": [],
        "deployment_slug": deployment_slug,
        "llm_provider": {
            "provider": "openai",
            "model": "intellirag-gpt-5.2",
            "api_key": "string",
            "base_url": "string"
        },
        "embedding_provider": {
            "provider": "openai",
            "model": "intellirag-text-embedding-3-small"
        },
        "llm_routing": "true",
        "system_prompt": "string",
        "summary_prompt": False,
        "followup_questions": False
    }
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {_AUTH_TOKEN}",
        "Content-Type": "application/json"
    }
    try:
        resp = http_requests.post(SEARCH_API_URL, json=payload, headers=headers, timeout=30)
        if resp.status_code != 200:
            return f"[API error {resp.status_code}]"
        data = resp.json()
        answer = data.get("answer") or ""
        if not answer:
            return "[No answer returned]"
        answer = re.sub(r'\[\d+\]', '', answer)
        answer = re.sub(r'^#{1,6}\s+', '', answer, flags=re.MULTILINE)
        return answer.strip()
    except http_requests.exceptions.Timeout:
        return "[timeout]"
    except Exception as exc:
        return f"[error: {exc}]"


def check_correctness(api_response: str, correct_answer: str) -> str:
    if not api_response or api_response.startswith("["):
        return "Manual Review"

    correct = (correct_answer or "").strip()
    response_lower = api_response.lower()

    # MCQ answer: (1), (2), (3), (4) — cannot reliably match free text
    if re.match(r'^\(\d+\)$', correct):
        return "Manual Review"

    try:
        correct_num = float(correct)
        found_nums = re.findall(r'-?\d+(?:\.\d+)?', api_response)
        for n in found_nums:
            if abs(float(n) - correct_num) < 0.01:
                return "Correct"
        return "Incorrect"
    except ValueError:
        pass

    if correct.lower() in response_lower:
        return "Correct"

    return "Manual Review"


# ── LaTeX → Unicode conversion ────────────────────────────────────────────────

_SUP = {
    '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
    '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹',
    '+': '⁺', '-': '⁻', '=': '⁼', '(': '⁽', ')': '⁾', 'n': 'ⁿ', 'i': 'ⁱ',
}

_HAT = {
    'i': 'î', 'j': 'ĵ', 'k': 'k̂',
    'x': 'x̂', 'y': 'ŷ', 'z': 'ẑ',
    'n': 'n̂', 'r': 'r̂',
}

_SUB = {
    '0': '₀', '1': '₁', '2': '₂', '3': '₃', '4': '₄',
    '5': '₅', '6': '₆', '7': '₇', '8': '₈', '9': '₉',
    '+': '₊', '-': '₋', '=': '₌', '(': '₍', ')': '₎',
    'a': 'ₐ', 'e': 'ₑ', 'o': 'ₒ', 'x': 'ₓ', 'h': 'ₕ', 'k': 'ₖ',
    'l': 'ₗ', 'm': 'ₘ', 'n': 'ₙ', 'p': 'ₚ', 's': 'ₛ', 't': 'ₜ',
    'i': 'ᵢ', 'r': 'ᵣ', 'u': 'ᵤ', 'v': 'ᵥ',
}

_GREEK = {
    'alpha': 'α', 'beta': 'β', 'gamma': 'γ', 'delta': 'δ',
    'epsilon': 'ε', 'varepsilon': 'ε', 'zeta': 'ζ', 'eta': 'η',
    'theta': 'θ', 'vartheta': 'ϑ', 'iota': 'ι', 'kappa': 'κ',
    'lambda': 'λ', 'mu': 'μ', 'nu': 'ν', 'xi': 'ξ',
    'pi': 'π', 'varpi': 'ϖ', 'rho': 'ρ', 'varrho': 'ϱ',
    'sigma': 'σ', 'varsigma': 'ς', 'tau': 'τ', 'upsilon': 'υ',
    'phi': 'φ', 'varphi': 'φ', 'chi': 'χ', 'psi': 'ψ', 'omega': 'ω',
    'Gamma': 'Γ', 'Delta': 'Δ', 'Theta': 'Θ', 'Lambda': 'Λ',
    'Xi': 'Ξ', 'Pi': 'Π', 'Sigma': 'Σ', 'Upsilon': 'Υ',
    'Phi': 'Φ', 'Psi': 'Ψ', 'Omega': 'Ω',
}

_SYMBOLS = {
    'times': '×', 'div': '÷', 'pm': '±', 'mp': '∓',
    'cdot': '·', 'cdots': '⋯', 'ldots': '…', 'dots': '…',
    'leq': '≤', 'le': '≤', 'geq': '≥', 'ge': '≥',
    'neq': '≠', 'ne': '≠', 'approx': '≈', 'equiv': '≡',
    'sim': '∼', 'simeq': '≃', 'cong': '≅', 'propto': '∝',
    'infty': '∞', 'partial': '∂', 'nabla': '∇',
    'int': '∫', 'iint': '∬', 'iiint': '∭', 'oint': '∮',
    'sum': '∑', 'prod': '∏',
    'forall': '∀', 'exists': '∃', 'nexists': '∄',
    'in': '∈', 'notin': '∉', 'ni': '∋',
    'subset': '⊂', 'supset': '⊃', 'subseteq': '⊆', 'supseteq': '⊇',
    'cup': '∪', 'cap': '∩', 'setminus': '∖', 'emptyset': '∅',
    'rightarrow': '→', 'to': '→', 'Rightarrow': '⇒',
    'leftarrow': '←', 'gets': '←', 'Leftarrow': '⇐',
    'leftrightarrow': '↔', 'Leftrightarrow': '⟺',
    'uparrow': '↑', 'downarrow': '↓', 'updownarrow': '↕',
    'longrightarrow': '⟶', 'longleftarrow': '⟵', 'mapsto': '↦',
    'angle': '∠', 'perp': '⊥', 'parallel': '∥',
    'circ': '∘', 'bullet': '•', 'therefore': '∴', 'because': '∵',
    'langle': '⟨', 'rangle': '⟩',
    'lfloor': '⌊', 'rfloor': '⌋', 'lceil': '⌈', 'rceil': '⌉',
    'oplus': '⊕', 'ominus': '⊖', 'otimes': '⊗', 'oslash': '⊘',
    'hbar': 'ℏ', 'ell': 'ℓ', 'Re': 'ℜ', 'Im': 'ℑ', 'aleph': 'ℵ',
    'vee': '∨', 'wedge': '∧', 'neg': '¬', 'lnot': '¬',
    'mid': '∣', 'nmid': '∤', 'll': '≪', 'gg': '≫',
    'triangle': '△', 'square': '□',
    ',': '', ';': ' ', ':': ' ', '!': '',
    'quad': '  ', 'qquad': '    ',
    'left': '', 'right': '',
    # Math function names — strip backslash, keep word as-is
    'sin': 'sin', 'cos': 'cos', 'tan': 'tan',
    'cot': 'cot', 'sec': 'sec', 'csc': 'csc',
    'arcsin': 'arcsin', 'arccos': 'arccos', 'arctan': 'arctan',
    'sinh': 'sinh', 'cosh': 'cosh', 'tanh': 'tanh', 'coth': 'coth',
    'log': 'log', 'ln': 'ln', 'exp': 'exp',
    'lim': 'lim', 'limsup': 'lim sup', 'liminf': 'lim inf',
    'max': 'max', 'min': 'min', 'sup': 'sup', 'inf': 'inf',
    'det': 'det', 'dim': 'dim', 'deg': 'deg',
    'gcd': 'gcd', 'lcm': 'lcm', 'mod': 'mod',
    'Pr': 'Pr', 'ker': 'ker', 'arg': 'arg',
}

_SKIP_CMDS = frozenset({
    'displaystyle', 'textstyle', 'scriptstyle', 'scriptscriptstyle',
    'rm', 'bf', 'it', 'sf', 'tt', 'cal', 'frak', 'bb',
    'limits', 'nolimits', 'nonumber',
    'big', 'Big', 'bigg', 'Bigg',
    'bigl', 'bigr', 'Bigl', 'Bigr', 'biggl', 'biggr', 'Biggl', 'Biggr',
})

_ARG_CMDS = frozenset({
    'text', 'mathrm', 'mathbf', 'mathit', 'mathsf', 'mathtt',
    'boldsymbol', 'overline', 'underline', 'widehat', 'widetilde',
    'overrightarrow', 'overleftarrow', 'overbrace', 'underbrace',
    'tilde', 'hat', 'bar', 'check', 'breve', 'acute', 'grave',
    'dot', 'ddot', 'dddot', 'vec',
    'mathcal', 'mathfrak', 'mathbb', 'operatorname',
})


def _extract_braced(s, pos):
    """Return (content, next_pos) for the {…} group starting at pos."""
    if pos >= len(s) or s[pos] != '{':
        return '', pos
    depth, start = 0, pos + 1
    for i in range(pos, len(s)):
        if s[i] == '{':
            depth += 1
        elif s[i] == '}':
            depth -= 1
            if depth == 0:
                return s[start:i], i + 1
    return s[start:], len(s)


def _cvt(expr):
    """Convert a LaTeX math expression (no delimiters) to Unicode."""
    out = []
    i = 0
    n = len(expr)
    while i < n:
        ch = expr[i]

        if ch == '\\':
            j = i + 1
            if j >= n:
                break
            if expr[j] == '\\':
                out.append('\n')
                i = j + 1
            elif expr[j] in '{}|.,':
                out.append(expr[j])
                i = j + 1
            elif expr[j].isalpha():
                k = j
                while k < n and expr[k].isalpha():
                    k += 1
                cmd = expr[j:k]
                i = k
                if i < n and expr[i] == ' ':
                    i += 1
                if cmd == 'frac':
                    num, i = _extract_braced(expr, i)
                    den, i = _extract_braced(expr, i)
                    out.append(f'{_cvt(num)}/{_cvt(den)}')
                elif cmd == 'sqrt':
                    if i < n and expr[i] == '[':
                        end = expr.find(']', i)
                        root = expr[i+1:end] if end != -1 else ''
                        i = end + 1 if end != -1 else i
                        content, i = _extract_braced(expr, i)
                        root_u = {'2': '²', '3': '³', '4': '⁴'}.get(root.strip(), root)
                        out.append(f'{root_u}√{_cvt(content)}')
                    elif i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                        out.append(f'√{_cvt(content)}')
                    else:
                        out.append('√')
                elif cmd in ('underset', 'overset'):
                    _, i = _extract_braced(expr, i)
                    base, i = _extract_braced(expr, i)
                    out.append(_cvt(base))
                elif cmd in ('overrightarrow', 'vec'):
                    if i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                        s = _cvt(content)
                    elif i < n and expr[i].isalpha():
                        s = expr[i]; i += 1
                    else:
                        s = ''
                    out.append(s + '⃗' if len(s) == 1 else f'({s})⃗')
                elif cmd == 'overleftarrow':
                    if i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                        s = _cvt(content)
                    elif i < n and expr[i].isalpha():
                        s = expr[i]; i += 1
                    else:
                        s = ''
                    out.append(s + '⃖' if len(s) == 1 else f'({s})⃖')
                elif cmd == 'hat':
                    if i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                    elif i < n and expr[i].isalpha():
                        content = expr[i]; i += 1
                    else:
                        content = ''
                    c = _cvt(content).strip()
                    out.append(_HAT.get(c, c + '̂'))
                elif cmd in _ARG_CMDS:
                    if i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                        out.append(_cvt(content))
                elif cmd in _SKIP_CMDS:
                    pass
                elif cmd in _GREEK:
                    out.append(_GREEK[cmd])
                elif cmd in _SYMBOLS:
                    out.append(_SYMBOLS[cmd])
                else:
                    if i < n and expr[i] == '{':
                        content, i = _extract_braced(expr, i)
                        out.append(_cvt(content))
            else:
                sym = _SYMBOLS.get(expr[j], expr[j])
                out.append(sym)
                i = j + 1

        elif ch == '^':
            i += 1
            if i < n and expr[i] == '{':
                content, i = _extract_braced(expr, i)
                out.append(''.join(_SUP.get(c, c) for c in _cvt(content)))
            elif i < n:
                out.append(_SUP.get(expr[i], expr[i]))
                i += 1

        elif ch == '_':
            i += 1
            if i < n and expr[i] == '{':
                content, i = _extract_braced(expr, i)
                out.append(''.join(_SUB.get(c, c) for c in _cvt(content)))
            elif i < n:
                out.append(_SUB.get(expr[i], expr[i]))
                i += 1

        elif ch in '{}':
            i += 1

        elif ch == '&':
            out.append('\t')
            i += 1

        else:
            out.append(ch)
            i += 1

    return ''.join(out)


def latex_to_unicode(text: str) -> str:
    """Replace LaTeX math and common markup in LLM output with Unicode equivalents.

    Handles: \\(...\\), \\[...\\], $$...$$, $...$, <sup>, <sub>, bare HTML tags.
    """
    if not text:
        return text
    # HTML sup/sub → Unicode
    text = re.sub(r'<sup>(.*?)</sup>',
                  lambda m: ''.join(_SUP.get(c, c) for c in m.group(1).strip()),
                  text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r'<sub>(.*?)</sub>',
                  lambda m: ''.join(_SUB.get(c, c) for c in m.group(1).strip()),
                  text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r'<[^>]+>', '', text)   # strip remaining HTML tags
    # \[...\] display math
    text = re.sub(r'\\\[(.+?)\\\]', lambda m: _cvt(m.group(1)), text, flags=re.DOTALL)
    # \(...\) inline math  (most common from vision-model output)
    text = re.sub(r'\\\((.+?)\\\)', lambda m: _cvt(m.group(1)), text, flags=re.DOTALL)
    # $$...$$ and $...$
    text = re.sub(r'\$\$(.+?)\$\$',  lambda m: _cvt(m.group(1)), text, flags=re.DOTALL)
    text = re.sub(r'\$([^$\n]+?)\$', lambda m: _cvt(m.group(1)), text)
    # Bare ^{...} / _{...} and ^n / _n outside any math delimiter.
    # Vision models frequently write answer-choice lines without delimiters,
    # so only the first (stem) line converts otherwise.
    def _sup_group(m):
        s = m.group(1)
        return ''.join(_SUP.get(c, c) for c in s) if all(c in _SUP for c in s) else f'^({s})'
    def _sub_group(m):
        s = m.group(1)
        return ''.join(_SUB.get(c, c) for c in s) if all(c in _SUB for c in s) else f'_({s})'
    text = re.sub(r'\^\{([^{}]+)\}', _sup_group, text)
    text = re.sub(r'_\{([^{}]+)\}',  _sub_group, text)
    # Multi-digit / signed numeric bare scripts: x^-12, H_20
    text = re.sub(r'\^(-?\d+)', lambda m: ''.join(_SUP.get(c, c) for c in m.group(1)), text)
    text = re.sub(r'_(-?\d+)',  lambda m: ''.join(_SUB.get(c, c) for c in m.group(1)), text)
    # Single-letter bare scripts: x^n, x_i, a_k — handle after multi-digit so ^12 isn't split
    text = re.sub(r'\^([a-zA-Z])', lambda m: _SUP.get(m.group(1), m.group(1)), text)
    text = re.sub(r'_([a-zA-Z])',  lambda m: _SUB.get(m.group(1), m.group(1)), text)
    return text


def build_validation_excel(results: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    has_figures = bool(results and "validated_figures" in results[0])

    has_reason  = bool(results and "reason" in results[0])
    header = ["Q #", "Excel Question", "PDF Answer", "Excel Answer", "Match", "Score %", "Status"]
    if has_reason:
        header.append("Reason")
    if has_figures:
        header += ["Image Figs", "Validated Figures", "Excel Figures", "Figs Match"]
    ws.append(header)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_styles = {
        "Correct":               (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                                  Font(color="006100", bold=True)),
        "Incorrect":             (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                                  Font(color="9C0006", bold=True)),
        "Manual Review":         (PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                                  Font(color="9C6500", bold=True)),
        "Not Found":             (PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                                  Font(color="595959", bold=True)),
        "Missing in Submission": (PatternFill(start_color="F2CEEF", end_color="F2CEEF", fill_type="solid"),
                                  Font(color="7030A0", bold=True)),
    }
    fig_match_styles = {
        True:  (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), Font(color="006100", bold=True)),
        False: (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), Font(color="9C0006", bold=True)),
    }

    for r in results:
        row_data = [
            r["q_num"],
            sanitize(r["excel_question"]),
            sanitize(r["pdf_answer"]),
            sanitize(r["excel_answer"]),
            r["match_type"],
            r["match_score"],
            r["status"],
        ]
        if has_reason:
            row_data.append(sanitize(r.get("reason", "")))
        if has_figures:
            fm = r.get("figures_match")
            row_data += [
                r.get("image_fig_count", 0),
                sanitize(r.get("validated_figures", "")),
                sanitize(r.get("excel_figures", "")),
                "N/A" if fm is None else ("Yes" if fm else "No"),
            ]
        ws.append(row_data)
        row = ws.max_row
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row, 2).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")

        fill, font = status_styles.get(r["status"], status_styles["Manual Review"])
        status_cell = ws.cell(row, 7)
        status_cell.fill = fill
        status_cell.font = font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        if has_reason:
            ws.cell(row, 8).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        if has_figures:
            fc = 8 + int(has_reason)   # column where "Image Figs" starts
            ws.cell(row, fc).alignment     = Alignment(horizontal="center", vertical="center")
            ws.cell(row, fc+1).alignment   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
            ws.cell(row, fc+2).alignment   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
            fm_cell = ws.cell(row, fc+3)
            fm_cell.alignment = Alignment(horizontal="center", vertical="center")
            fm = r.get("figures_match")
            if fm is not None:
                fm_fill, fm_font = fig_match_styles[fm]
                fm_cell.fill = fm_fill
                fm_cell.font = fm_font

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22
    if has_reason:
        ws.column_dimensions['H'].width = 50
    if has_figures:
        fc = 8 + int(has_reason)
        ws.column_dimensions[get_column_letter(fc)].width   = 12
        ws.column_dimensions[get_column_letter(fc+1)].width = 30
        ws.column_dimensions[get_column_letter(fc+2)].width = 30
        ws.column_dimensions[get_column_letter(fc+3)].width = 12

    wb.save(output_path)


def build_evaluation_excel(
    questions: list,
    answers: list,
    api_responses: list,
    statuses: list,
    output_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    ws.append(["Question #", "Question", "Correct Answer", "API Response", "Status"])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    status_styles = {
        "Correct":       (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                          Font(color="006100", bold=True)),
        "Incorrect":     (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                          Font(color="9C0006", bold=True)),
        "Manual Review": (PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                          Font(color="9C6500", bold=True)),
    }

    for idx, (question, answer, api_resp, status) in enumerate(
        zip(questions, answers, api_responses, statuses), start=1
    ):
        ws.append([idx, sanitize(question), sanitize(answer), sanitize(api_resp), status])
        row = ws.max_row
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        fill, font = status_styles.get(status, status_styles["Manual Review"])
        status_cell = ws.cell(row, 5)
        status_cell.fill = fill
        status_cell.font = font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 18

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
