"""
PDF Processor Module
Extracts questions and answers from PDF documents and exports to Excel.
"""

import os
import re
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

try:
    import fitz  # PyMuPDF
    _PYMUPDF_AVAILABLE = True
except ImportError:
    _PYMUPDF_AVAILABLE = False

try:
    import pdfplumber
    _PDFPLUMBER_AVAILABLE = True
except ImportError:
    _PDFPLUMBER_AVAILABLE = False



class PDFProcessor:
    """
    Processes PDF documents containing questions and answers.
    Extracts content and exports to Excel format.
    """

    def __init__(self, questions_pdf_path: str, answers_pdf_path: str):
        """
        Initialize the PDF processor with question and answer PDF paths.
        
        Args:
            questions_pdf_path: Path to the PDF containing questions
            answers_pdf_path: Path to the PDF containing answers (correct options)
        """
        self.questions_pdf_path = questions_pdf_path
        self.answers_pdf_path = answers_pdf_path
        self.questions = []
        self.answers = []

    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """
        Extract all text from a PDF file.

        Uses PyMuPDF (best Unicode/math symbol support) → pdfplumber → PyPDF2
        as a fallback chain so that Greek letters, mathematical operators,
        subscripts, superscripts, and chemical formulas are preserved.

        Args:
            pdf_path: Path to the PDF file

        Returns:
            Extracted text from the PDF

        Raises:
            FileNotFoundError: If PDF file doesn't exist
            Exception: If PDF reading fails
        """
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")

        # --- PyMuPDF (best for math/chemistry notation) ---
        if _PYMUPDF_AVAILABLE:
            try:
                doc = fitz.open(pdf_path)
                pages_text = []
                for page in doc:
                    pages_text.append(page.get_text("text"))
                doc.close()
                text = "\n".join(pages_text)
                if text.strip():
                    return text
            except Exception:
                pass  # fall through to next extractor

        # --- pdfplumber (good ligature/layout handling) ---
        if _PDFPLUMBER_AVAILABLE:
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    pages_text = []
                    for page in pdf.pages:
                        extracted = page.extract_text()
                        if extracted:
                            pages_text.append(extracted)
                text = "\n".join(pages_text)
                if text.strip():
                    return text
            except Exception:
                pass

        raise Exception(
            f"No PDF extraction library available. "
            f"Install pymupdf or pdfplumber."
        )

    def parse_questions(self, text: str) -> List[str]:
        """
        Parse questions from extracted text.
        Supports JEE Main format: Q{n}. JEE Main {year} (date Shift n)
        Falls back to numbered (1., 2., ...) or Q: prefix formats.

        Args:
            text: Raw text extracted from PDF

        Returns:
            List of parsed questions in question-number order
        """
        # JEE Main format: Q1. JEE Main 2025 (8 April Shift 2) ...
        jee_pattern = r'Q(\d+)\.\s+(.*?)(?=\nQ\d+\.|$)'
        matches = re.findall(jee_pattern, text, re.DOTALL)

        if matches:
            parsed = sorted([(int(num), content.strip()) for num, content in matches])
            return [content for _, content in parsed]

        # Standard numbered format: 1. question text
        numbered_pattern = r'^\s*\d+\.\s+(.+?)(?=\n\s*\d+\.|$)'
        matches = re.findall(numbered_pattern, text, re.MULTILINE | re.DOTALL)
        if matches:
            return [q.strip() for q in matches]

        # Q: prefix format
        lines = text.split('\n')
        return [line.strip() for line in lines if line.strip().startswith('Q:')]

    def parse_answers(self, text: str) -> List[str]:
        """
        Parse answers from extracted text.
        Supports JEE Main answer key format: {n}. (k) or {n}. number
        Examples:
            1. (2)    → MCQ option 2
            2. 12     → numerical answer 12
            134. 2.18 → decimal answer 2.18
        Falls back to 'Answer: X' or single-letter A/B/C/D formats.

        Args:
            text: Raw text extracted from PDF

        Returns:
            List of answers indexed by question number (index 0 = Q1's answer)
        """
        # JEE Main answer key: number. followed by (k) or a plain/decimal number
        jee_pattern = r'\b(\d+)\.\s+(\(\d+\)|\d+(?:\.\d+)?)'
        matches = re.findall(jee_pattern, text)

        if matches:
            answer_dict = {}
            for num_str, answer in matches:
                answer_dict[int(num_str)] = answer
            if answer_dict:
                max_num = max(answer_dict.keys())
                return [answer_dict.get(i, "N/A") for i in range(1, max_num + 1)]

        # Standard "Answer: X" format
        answer_pattern = r'Answer:\s*([A-D])'
        matches = re.findall(answer_pattern, text, re.IGNORECASE)
        if matches:
            return matches

        # Single letter answers
        lines = text.split('\n')
        return [line.strip() for line in lines if line.strip() in ['A', 'B', 'C', 'D']]

    def process_and_export(self, output_excel_path: str) -> str:
        """
        Process PDFs and export questions with answers to Excel.
        
        Args:
            output_excel_path: Path where the Excel file will be saved
            
        Returns:
            Path to the created Excel file
            
        Raises:
            Exception: If processing fails
        """
        try:
            # Extract text from both PDFs
            questions_text = self.extract_text_from_pdf(self.questions_pdf_path)
            answers_text = self.extract_text_from_pdf(self.answers_pdf_path)
            
            # Parse content
            self.questions = self.parse_questions(questions_text)
            self.answers = self.parse_answers(answers_text)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Q&A"
            
            # Add headers
            headers = ["Question #", "Question", "Correct Answer"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add data rows
            for idx, question in enumerate(self.questions, 1):
                answer = self.answers[idx - 1] if idx - 1 < len(self.answers) else "N/A"
                ws.append([idx, question, answer])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 18
            
            # Center align all cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                row[0].alignment = Alignment(horizontal="center", vertical="top")
                row[2].alignment = Alignment(horizontal="center", vertical="center")
                row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Save workbook
            os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
            wb.save(output_excel_path)
            
            print(f"✓ Successfully exported {len(self.questions)} questions to {output_excel_path}")
            return output_excel_path
            
        except Exception as e:
            print(f"✗ Error processing PDFs: {str(e)}")
            raise

    def get_summary(self) -> Dict:
        """
        Get summary of extracted questions and answers.
        
        Returns:
            Dictionary with summary statistics
        """
        return {
            "total_questions": len(self.questions),
            "total_answers": len(self.answers),
            "matched_pairs": min(len(self.questions), len(self.answers)),
            "timestamp": datetime.now().isoformat()
        }
