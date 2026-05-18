"""
Non-Flask helpers for the extraction pipeline.
All logic previously embedded in api.py lives here so that api.py contains
only route definitions.
"""
import os
import re
import base64
import json
from concurrent.futures import ThreadPoolExecutor

import fitz as _fitz
import requests as _http
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.pdf_utils import (
    crop_questions_from_pdf,
    extract_figures_from_pdf,
    build_question_mapping,
    detect_layout_fitz,
    extract_figures_per_question,
)
from src.vision import call_vision, call_vision_with_prompt, _MODEL_ALIASES
from src.mathpix import call_mathpix
from src.page_classifier import classify_page_with_gpt
from src.helpers import sanitize, latex_to_unicode

_VISION_MAX_WORKERS = 8

_VLM_VALIDATE_URL   = "http://localhost:11434/api/chat"
_VLM_VALIDATE_MODEL = "qwen2.5vl:7b"


# ── Prompts ────────────────────────────────────────────────────────────────────

QUESTION_TRANSCRIBE_PROMPT = """\
This image shows exactly one exam question. Extract it completely.

Rules:
1. Include the question number, the full question stem, and ALL answer choices \
   (e.g. (1)(2)(3)(4) or (A)(B)(C)(D)) exactly as printed.
2. If a figure, diagram, graph, or image appears anywhere in the question, \
   insert [Figure 1], [Figure 2], etc. at that exact position in the text.
3. Write mathematics in plain Unicode — fractions as a/b, square roots as √x, \
   powers as x² — no LaTeX, no backslashes.
4. Ignore watermarks, page numbers, footers, and any text outside this question.
5. Output ONLY the extracted question text, nothing else.
"""

SOLUTION_CROP_PROMPT = """\
This image shows a solution or answer entry from an exam paper.
Extract ONLY the FIRST (topmost) solution visible in the image.

Rules:
1. Include the question number and the full answer/solution text exactly as printed.
2. If the answer is a single option (e.g. (a), (b), (c), (d)), output just that option.
3. If it is a worked solution, include all steps.
4. If multiple partial entries appear (e.g. a second question number begins near the \
   bottom), extract ONLY the entry whose number appears at the TOP — stop before \
   any subsequent question number.
5. Write mathematics in plain Unicode — fractions as a/b, square roots as √x, \
   powers as x² — no LaTeX, no backslashes.
6. Ignore watermarks, page numbers, or any text outside the top solution entry.
7. Output ONLY the extracted answer text, nothing else.
"""

_VLM_COMPARE_PROMPT = (
    "You are a precise exam-question validator.\n\n"
    "The image shows a question cropped from the original exam PDF.\n"
    "Below is the text that was transcribed for this question:\n\n"
    "TRANSCRIPTION:\n{excel_text}\n\n"
    "FIGURE PLACEHOLDERS: The transcription may contain tokens like [Figure 1], "
    "[Figure 2], etc. Each token represents an embedded visual element (figure, "
    "diagram, graph, or image-based answer option) at that position in the question. "
    "A placeholder is correct if a visual element appears at the corresponding "
    "position in the image, and the numbering follows reading order (top-to-bottom).\n\n"
    "Decide whether the transcription is an accurate and complete representation "
    "of the question in the image.\n\n"
    "Evaluate:\n"
    "1. Is the question stem word-for-word correct (wording, numbers, math, units)?\n"
    "2. Are all answer choices present and correctly transcribed?\n"
    "3. Is mathematical notation (fractions, exponents, symbols) accurately captured?\n"
    "4. Are [Figure N] placeholders present wherever a visual appears, in the right positions?\n\n"
    "If the transcription is not a perfect match, list each specific discrepancy in the "
    "'issues' array. Each issue must be concrete and quote the conflicting text.\n\n"
    'Return ONLY valid JSON with no surrounding text:\n'
    '{{"match": true/false, "issues": ["..."], "confidence": 0.0-1.0, "figure_count": N}}\n'
    'where figure_count is the number of distinct visual elements visible in the image.'
)


# ── Regex ──────────────────────────────────────────────────────────────────────

# Matches compact answer-key entries: "1.(a)", "2. (c)", "3.b", "10) d", "18.(14.00)"
# Group 1 = question number, Group 2 = answer (letter a-d/A-D OR numeric value)
_ANS_KEY_ENTRY_RE = re.compile(
    r'(?<!\d)(\d{1,3})\s*[.)]\s*\(?\s*([a-dA-D]|\d+(?:\.\d+)?)\s*\)?(?!\w)',
)


# ── Page classification ────────────────────────────────────────────────────────

def classify_pages(pdf_path: str, page_images: list) -> list:
    """Classify each page and detect its column layout.

    Returns a list of dicts with keys:
        page        — 1-based page number
        page_type   — theory | questions | solutions | misc
        layout_type — single_column | multi_column
        confidence, reason
    """
    results = []
    for page_num, image_path in enumerate(page_images, start=1):
        classification = classify_page_with_gpt(image_path)
        page_type = classification.get("page_type", "misc")
        print(
            f"[PAGE {page_num:03d}] {page_type.upper():<12} "
            f"conf={classification.get('confidence', 0):.2f}  "
            f"{classification.get('reason', '')}"
        )

        layout_type = "single_column"
        if page_type in ("questions", "solutions"):
            layout = detect_layout_fitz(pdf_path, page_num - 1)
            layout_type = layout.get("layout", "single_column")

        results.append({
            "page":        page_num,
            "page_type":   page_type,
            "layout_type": layout_type,
            "confidence":  classification.get("confidence"),
            "reason":      classification.get("reason"),
        })
    return results


# ── VLM transcription ──────────────────────────────────────────────────────────

def transcribe_crops(crops: dict, prompt: str, model: str) -> dict:
    """VLM-transcribe a {q_num: image_path} dict in parallel.

    Returns {q_num: text} for every crop that produced non-empty output.
    """
    if not crops:
        return {}

    resolved = _MODEL_ALIASES.get(model, model)
    workers = 3 if resolved.startswith("claude") else _VISION_MAX_WORKERS

    def _do(item):
        q_num, crop_path = item
        try:
            text = call_vision_with_prompt(crop_path, prompt, model)
            return q_num, text.strip()
        except Exception:
            return q_num, ""

    texts = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for q_num, text in pool.map(_do, crops.items()):
            if text:
                texts[q_num] = text
    return texts


# ── Answer key extraction ──────────────────────────────────────────────────────

def extract_answer_keys_from_regions(pdf_path: str, s_indices: list):
    """
    Detect compact answer-key regions on hybrid solution pages.

    Returns:
        answers: {global_q_num: "(a)"}
        answer_regions: {page_idx: (y0, y1)}
            vertical regions occupied by answer keys
            so crop_solutions_from_pages can ignore them
    """

    doc = _fitz.open(pdf_path)

    answers = {}
    answer_regions = {}

    expected_num = None
    section_offset = 0

    try:
        for page_idx in sorted(s_indices):

            if page_idx >= len(doc):
                continue

            page = doc[page_idx]

            blocks = page.get_text("blocks")

            answer_lines = []

            in_answers_section = False
            region_y0 = None
            region_y1 = None

            for block in sorted(blocks, key=lambda b: (b[1], b[0])):

                x0, y0, x1, y1, text, *_ = block

                clean = re.sub(r"\s+", " ", text.strip())

                # Start of answer section
                if re.search(r"\bAnswers\b", clean, re.I):
                    in_answers_section = True
                    region_y0 = y0
                    continue

                # End of answer section
                if re.search(r"\bSolutions\b", clean, re.I):
                    region_y1 = y0
                    break

                if not in_answers_section:
                    continue

                answer_lines.append(clean)

            # No hybrid answer section
            if not answer_lines:
                continue

            # Store exclusion region
            if region_y0 is not None and region_y1 is not None:
                answer_regions[page_idx] = (
                    max(0, region_y0 - 10),
                    region_y1 + 5
                )

            # Parse compact entries
            text_blob = " ".join(answer_lines)

            matches = re.findall(
                r"(\d+)\s*\.\s*\(([a-zA-Z0-9\.\-]+)\)",
                text_blob
            )

            for q_str, ans in matches:

                num = int(q_str)

                # Section rollover detection
                if expected_num is None or num >= expected_num:
                    global_num = num + section_offset
                    expected_num = num + 1

                elif (
                    num < expected_num
                    and num <= 5
                    and (expected_num - num) > 5
                ):
                    section_offset += expected_num - 1
                    global_num = num + section_offset
                    expected_num = num + 1

                else:
                    continue

                answers[global_num] = f"({ans.lower()})"

    finally:
        doc.close()

    return answers, answer_regions

# ── Excel writers ──────────────────────────────────────────────────────────────

def write_questions_excel(result: list, output_path: str) -> None:
    """Write rows with columns: question_num, question_text, figures, answers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    ws.append(["question_num", "question_text", "figures", "answers"])
    _apply_header_style(ws)

    for entry in result:
        ws.append([entry["question_num"], entry["question_text"],
                   entry["figures"], entry["answers"]])
        row = ws.max_row
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row, 2).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    wb.save(output_path)


def write_general_extraction_excel(result: list, output_path: str) -> None:
    """Write rows with columns: question_num, question_text, figures, answer_key, solution."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    ws.append(["question_num", "question_text", "figures", "answer_key", "solution"])
    _apply_header_style(ws)

    for entry in result:
        ws.append([
            entry["question_num"],
            entry["question_text"],
            entry["figures"],
            entry["answer_key"],
            entry["solution"],
        ])
        row = ws.max_row
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row, 2).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 5).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 70
    wb.save(output_path)


def _apply_header_style(ws) -> None:
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Old-pipeline helpers (used by /api/pdf-to-images and /api/extract-mathpix) ─

def prepare_work_dirs(base_dir: str) -> tuple:
    questions_dir = os.path.join(base_dir, "questions")
    figures_dir   = os.path.join(base_dir, "figures")
    os.makedirs(questions_dir, exist_ok=True)
    os.makedirs(figures_dir, exist_ok=True)
    return questions_dir, figures_dir


def run_pdf_pipeline(questions_path: str, answers_path: str,
                     questions_dir: str, figures_dir: str) -> tuple:
    crop_by_qnum = crop_questions_from_pdf(questions_path, questions_dir)
    fig_data     = extract_figures_from_pdf(questions_path, figures_dir)
    mapping      = build_question_mapping(questions_path, answers_path, fig_data)
    return crop_by_qnum, mapping


def transcribe_entry(entry: dict, crop_by_qnum: dict, model: str) -> dict:
    crop_path = crop_by_qnum.get(entry["question_num"])
    figs = entry.get("figure") or []
    if crop_path and os.path.exists(crop_path):
        try:
            q_text = call_vision(crop_path, figure_count=len(figs), model=model)
        except Exception as exc:
            q_text = f"[vision error: {exc}]"
    else:
        q_text = ""
    return {
        "question_num":  str(entry["question_num"]),
        "question_text": sanitize(latex_to_unicode(q_text)),
        "answers":       sanitize(latex_to_unicode(entry.get("answer", "N/A") or "N/A")),
        "figures":       ", ".join(os.path.basename(p) for p in figs),
    }


def transcribe_all_parallel(mapping: list, crop_by_qnum: dict, model: str) -> list:
    workers = 3 if model.startswith("claude") else _VISION_MAX_WORKERS
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(transcribe_entry, entry, crop_by_qnum, model)
                   for entry in mapping]
    return [f.result() for f in futures]


def transcribe_entry_mathpix(entry: dict, crop_by_qnum: dict, model: str) -> dict:
    crop_path = crop_by_qnum.get(entry["question_num"])
    figs = entry.get("figure") or []
    if crop_path and os.path.exists(crop_path):
        try:
            q_text = call_mathpix(crop_path, model=model)
        except Exception as exc:
            q_text = f"[mathpix error: {exc}]"
    else:
        q_text = ""
    return {
        "question_num":  str(entry["question_num"]),
        "question_text": sanitize(latex_to_unicode(q_text)),
        "answers":       sanitize(latex_to_unicode(entry.get("answer", "N/A") or "N/A")),
        "figures":       ", ".join(os.path.basename(p) for p in figs),
    }


def transcribe_all_mathpix_parallel(mapping: list, crop_by_qnum: dict, model: str) -> list:
    with ThreadPoolExecutor(max_workers=_VISION_MAX_WORKERS) as executor:
        futures = [executor.submit(transcribe_entry_mathpix, entry, crop_by_qnum, model)
                   for entry in mapping]
    return [f.result() for f in futures]


# ── Validate-route helpers ─────────────────────────────────────────────────────

def vlm_compare_question(image_path: str, excel_text: str) -> dict:
    """Send a PDF question crop + Excel transcription to a local VLM for comparison."""
    prompt = _VLM_COMPARE_PROMPT.format(excel_text=excel_text.strip() or "(empty)")
    try:
        with open(image_path, "rb") as fh:
            image_b64 = base64.b64encode(fh.read()).decode()
        payload = {
            "model": _VLM_VALIDATE_MODEL,
            "messages": [{"role": "user", "content": prompt, "images": [image_b64]}],
            "stream": False,
            "options": {"temperature": 0, "num_predict": 512},
        }
        resp = _http.post(_VLM_VALIDATE_URL, json=payload, timeout=120)
        resp.raise_for_status()
        raw = resp.json()["message"]["content"].strip()
        return json.loads(raw[raw.find("{"):raw.rfind("}") + 1])
    except Exception as exc:
        return {"match": False, "issues": [f"VLM error: {exc}"], "confidence": 0.0, "error": True}


def normalise_cols(df) -> dict:
    """Return {normalised_name: original_column_name} for all DataFrame columns."""
    return {
        c.strip().lower().replace(" ", "_").replace("#", "num"): c
        for c in df.columns
    }


def pick_col(norm: dict, aliases: list):
    for alias in aliases:
        if alias in norm:
            return norm[alias]
    return None
