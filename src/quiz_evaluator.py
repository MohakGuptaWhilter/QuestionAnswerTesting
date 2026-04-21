"""
Quiz Evaluator Module
Tests answers via API endpoint, validates correctness, and generates reports.
"""

import os
from typing import List, Dict, Optional
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from enum import Enum


class ResultStatus(Enum):
    """Status of a quiz answer result."""
    CORRECT = "Correct"
    INCORRECT = "Incorrect"
    ERROR = "Error"
    SKIPPED = "Skipped"


class QuizEvaluator:
    """
    Evaluates quiz answers by calling an API endpoint and validates correctness.
    Generates detailed reports in Excel format.
    """

    def __init__(self, excel_file_path: str, api_endpoint: str, api_timeout: int = 10):
        """
        Initialize the quiz evaluator.
        
        Args:
            excel_file_path: Path to Excel file containing questions
            api_endpoint: API endpoint URL to submit answers for validation
            api_timeout: Timeout for API requests in seconds
        """
        self.excel_file_path = excel_file_path
        self.api_endpoint = api_endpoint
        self.api_timeout = api_timeout
        self.results = []
        self.questions = []

    def load_questions_from_excel(self) -> List[Dict]:
        """
        Load questions from Excel file.
        
        Returns:
            List of dictionaries containing question data
            
        Raises:
            FileNotFoundError: If Excel file doesn't exist
        """
        if not os.path.exists(self.excel_file_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_file_path}")

        try:
            wb = load_workbook(self.excel_file_path)
            ws = wb.active
            
            questions = []
            # Skip header row (row 1)
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value is not None:
                    question_data = {
                        "id": row[0].value,
                        "question": row[1].value,
                        "correct_answer": row[2].value
                    }
                    questions.append(question_data)
            
            self.questions = questions
            return questions
            
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")

    def validate_answer_with_api(self, question_id: int, user_answer: str) -> Dict:
        """
        Validate an answer by calling the API endpoint.
        
        Args:
            question_id: ID of the question
            user_answer: The answer provided by the user
            
        Returns:
            Dictionary containing validation result
        """
        try:
            payload = {
                "question_id": question_id,
                "user_answer": user_answer
            }
            
            response = requests.post(
                self.api_endpoint,
                json=payload,
                timeout=self.api_timeout
            )
            
            if response.status_code == 200:
                api_result = response.json()
                return {
                    "status": ResultStatus.CORRECT if api_result.get("correct", False) else ResultStatus.INCORRECT,
                    "is_correct": api_result.get("correct", False),
                    "expected_answer": api_result.get("correct_answer"),
                    "error": None
                }
            else:
                return {
                    "status": ResultStatus.ERROR,
                    "is_correct": False,
                    "expected_answer": None,
                    "error": f"API returned status {response.status_code}"
                }
                
        except requests.exceptions.Timeout:
            return {
                "status": ResultStatus.ERROR,
                "is_correct": False,
                "expected_answer": None,
                "error": "API request timeout"
            }
        except requests.exceptions.ConnectionError:
            return {
                "status": ResultStatus.ERROR,
                "is_correct": False,
                "expected_answer": None,
                "error": "Could not connect to API"
            }
        except Exception as e:
            return {
                "status": ResultStatus.ERROR,
                "is_correct": False,
                "expected_answer": None,
                "error": str(e)
            }

    def test_answers(self, user_answers: List[str]) -> List[Dict]:
        """
        Test user answers against API and collect results.
        
        Args:
            user_answers: List of answers provided by the user
                         (should match order of questions in Excel)
            
        Returns:
            List of result dictionaries
        """
        if not self.questions:
            self.load_questions_from_excel()
        
        results = []
        
        for idx, question in enumerate(self.questions):
            user_answer = user_answers[idx] if idx < len(user_answers) else None
            
            if user_answer is None:
                result = {
                    "question_id": question["id"],
                    "question": question["question"],
                    "user_answer": None,
                    "status": ResultStatus.SKIPPED,
                    "is_correct": False,
                    "expected_answer": question["correct_answer"],
                    "error": "No answer provided"
                }
            else:
                # Validate answer with API
                api_response = self.validate_answer_with_api(
                    question["id"],
                    user_answer
                )
                
                result = {
                    "question_id": question["id"],
                    "question": question["question"],
                    "user_answer": user_answer,
                    "status": api_response["status"],
                    "is_correct": api_response["is_correct"],
                    "expected_answer": api_response.get("expected_answer", question["correct_answer"]),
                    "error": api_response.get("error")
                }
            
            results.append(result)
        
        self.results = results
        return results

    def export_results_to_excel(self, output_excel_path: str) -> str:
        """
        Export test results to Excel file with detailed formatting.
        
        Args:
            output_excel_path: Path where results Excel file will be saved
            
        Returns:
            Path to the created Excel file
        """
        try:
            wb = load_workbook(self.excel_file_path)
            ws = wb.active
            
            # Add new columns for results
            ws['D1'] = "User Answer"
            ws['E1'] = "Result"
            ws['F1'] = "Expected Answer"
            ws['G1'] = "Error"
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in ['D', 'E', 'F', 'G']:
                cell = ws[f'{col}1']
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add results data
            for row_idx, result in enumerate(self.results, 2):
                ws[f'D{row_idx}'] = result.get("user_answer", "N/A")
                ws[f'E{row_idx}'] = result["status"].value
                ws[f'F{row_idx}'] = result.get("expected_answer", "N/A")
                ws[f'G{row_idx}'] = result.get("error", "")
                
                # Color code results
                if result["status"] == ResultStatus.CORRECT:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    font = Font(color="006100", bold=True)
                elif result["status"] == ResultStatus.INCORRECT:
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    font = Font(color="9C0006", bold=True)
                else:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    font = Font(color="9C6500")
                
                ws[f'E{row_idx}'].fill = fill
                ws[f'E{row_idx}'].font = font
                ws[f'E{row_idx}'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Adjust column widths
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 30
            
            # Save results workbook
            os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
            wb.save(output_excel_path)
            
            # Print summary
            correct_count = sum(1 for r in self.results if r["is_correct"])
            total_count = len(self.results)
            percentage = (correct_count / total_count * 100) if total_count > 0 else 0
            
            print(f"\n{'='*50}")
            print(f"Quiz Test Results")
            print(f"{'='*50}")
            print(f"Total Questions: {total_count}")
            print(f"Correct Answers: {correct_count}")
            print(f"Incorrect Answers: {total_count - correct_count}")
            print(f"Accuracy: {percentage:.1f}%")
            print(f"Results saved to: {output_excel_path}")
            print(f"{'='*50}\n")
            
            return output_excel_path
            
        except Exception as e:
            print(f"✗ Error exporting results: {str(e)}")
            raise

    def get_summary(self) -> Dict:
        """
        Get summary statistics of the test results.
        
        Returns:
            Dictionary with summary statistics
        """
        if not self.results:
            return {
                "total_questions": 0,
                "correct_answers": 0,
                "incorrect_answers": 0,
                "errors": 0,
                "accuracy": 0.0
            }
        
        correct = sum(1 for r in self.results if r["is_correct"])
        incorrect = sum(1 for r in self.results if r["status"] == ResultStatus.INCORRECT)
        errors = sum(1 for r in self.results if r["status"] == ResultStatus.ERROR)
        total = len(self.results)
        
        return {
            "total_questions": total,
            "correct_answers": correct,
            "incorrect_answers": incorrect,
            "errors": errors,
            "accuracy": (correct / total * 100) if total > 0 else 0.0,
            "timestamp": datetime.now().isoformat()
        }
