"""
Flask API — endpoint definitions only.
All helper logic (prompts, cropping orchestration, Excel writers, etc.) lives in
src/pipeline_helpers.py.
"""
import os
import tempfile
import shutil

from flask import Flask, request, send_file, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from rapidfuzz import fuzz as _fuzz

from arihant_pipeline.exporter import Exporter
from arihant_pipeline.general_purpose_extraction_pipeline import GeneralPurposeExtractionPipeline
from arihant_pipeline.pipeline_config import PipelineConfig

from src.pdf_processor import PDFProcessor
from src.quickstart import parse_pdf
from src.pdf_utils import (
    pdf_pages_to_png,
    crop_questions_from_pages,
    crop_solutions_from_pages,
    crop_questions_from_pdf,
    extract_figures_from_pages, map_figures_to_questions_on_pages,
    extract_figures_per_question,
)
from src.helpers import (
    sanitize, latex_to_unicode,
    check_correctness, build_validation_excel,
    FIGURE_URL_RE, FIG_FONT, inline_fig_labels,
)
from src.pipeline_helpers import (
    QUESTION_TRANSCRIBE_PROMPT, SOLUTION_CROP_PROMPT,
    classify_pages, extract_answer_keys_from_regions, transcribe_crops,
    write_questions_excel, write_general_extraction_excel,
    prepare_work_dirs, run_pdf_pipeline,
    transcribe_all_parallel, transcribe_all_mathpix_parallel,
    vlm_compare_question, normalise_cols, pick_col,
)


app = Flask(__name__)

UPLOAD_FOLDER = tempfile.gettempdir()
ALLOWED_EXTENSIONS = {'pdf'}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500 MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE


# ── Request helpers ────────────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_request():
    if 'questions_pdf' not in request.files or 'answers_pdf' not in request.files:
        return False, "Missing required files: 'questions_pdf' and 'answers_pdf'"
    q = request.files['questions_pdf']
    a = request.files['answers_pdf']
    if q.filename == '' or a.filename == '':
        return False, "File names cannot be empty"
    if not (allowed_file(q.filename) and allowed_file(a.filename)):
        return False, "Only PDF files are allowed"
    return True, None


# ── Error handlers ─────────────────────────────────────────────────────────────

@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({"error": f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB"}), 413

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Endpoint not found"}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Internal server error"}), 500


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "healthy", "service": "QA-PDF-Extractor-API", "version": "1.0.0"}), 200


@app.route('/api/extract', methods=['POST'])
def extract_qa():
    questions_path = answers_path = None
    try:
        is_valid, error_msg = validate_request()
        if not is_valid:
            return jsonify({"error": error_msg}), 400

        questions_file = request.files['questions_pdf']
        answers_file   = request.files['answers_pdf']
        questions_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(questions_file.filename))
        answers_path   = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(answers_file.filename))
        questions_file.save(questions_path)
        answers_file.save(answers_path)

        questions_md   = parse_pdf(questions_path)["markdown"]
        processor      = PDFProcessor(questions_path, answers_path)
        questions_list = processor.parse_questions(questions_md)
        answers_list   = processor.parse_answers(processor.extract_text_from_pdf(answers_path))

        if not questions_list:
            return jsonify({"error": "No questions could be extracted from the PDF"}), 422

        wb = Workbook()
        ws = wb.active
        ws.title = "Q&A"
        max_figs = max((len(FIGURE_URL_RE.findall(q)) for q in questions_list), default=0)
        ans_col  = 3 + max_figs
        header   = ["Question #", "Question"] + [f"Figure {n}" for n in range(1, max_figs + 1)] + ["Correct Answer"]
        ws.append(header)
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, question in enumerate(questions_list, start=1):
            answer = answers_list[idx - 1] if idx - 1 < len(answers_list) else "N/A"
            urls   = FIGURE_URL_RE.findall(question)
            q_text = latex_to_unicode(sanitize(inline_fig_labels(question)))
            ws.append([idx, q_text] + [None] * max_figs + [sanitize(answer)])
            row = ws.max_row
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, 2).alignment = Alignment(horizontal="left",   vertical="top", wrap_text=True)
            for n, url in enumerate(urls):
                fig_cell           = ws.cell(row, 3 + n)
                fig_cell.value     = f"View Figure {n + 1}"
                fig_cell.hyperlink = url
                fig_cell.font      = FIG_FONT
                fig_cell.alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, ans_col).alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        for n in range(max_figs):
            ws.column_dimensions[get_column_letter(3 + n)].width = 15
        ws.column_dimensions[get_column_letter(ans_col)].width = 18

        output_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'extracted_qa.xlsx')
        wb.save(output_excel)
        return send_file(output_excel, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'qa_extract_{len(questions_list)}q.xlsx')

    except FileNotFoundError as e:
        return jsonify({"error": f"File not found: {str(e)}"}), 404
    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        for path in (questions_path, answers_path):
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@app.route('/api/pdf-to-images', methods=['POST'])
def pdf_to_images():
    questions_path = answers_path = None
    try:
        is_valid, error_msg = validate_request()
        if not is_valid:
            return jsonify({"error": error_msg}), 400

        model          = request.form.get("model", "qwen2.5vl:7b")
        questions_file = request.files['questions_pdf']
        answers_file   = request.files['answers_pdf']
        questions_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(questions_file.filename))
        answers_path   = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(answers_file.filename))
        questions_file.save(questions_path)
        answers_file.save(answers_path)

        questions_dir, figures_dir = prepare_work_dirs(os.getcwd())
        crop_by_qnum, mapping = run_pdf_pipeline(questions_path, answers_path, questions_dir, figures_dir)
        result = transcribe_all_parallel(mapping, crop_by_qnum, model)

        output_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'questions_output.xlsx')
        write_questions_excel(result, output_excel)
        return send_file(output_excel, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='questions_output.xlsx')

    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        for path in (questions_path, answers_path):
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@app.route('/api/extract-mathpix', methods=['POST'])
def extract_mathpix():
    questions_path = answers_path = None
    try:
        is_valid, error_msg = validate_request()
        if not is_valid:
            return jsonify({"error": error_msg}), 400

        model          = request.form.get("model", "text")
        questions_file = request.files['questions_pdf']
        answers_file   = request.files['answers_pdf']
        questions_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(questions_file.filename))
        answers_path   = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(answers_file.filename))
        questions_file.save(questions_path)
        answers_file.save(answers_path)

        questions_dir, figures_dir = prepare_work_dirs(os.getcwd())
        crop_by_qnum, mapping = run_pdf_pipeline(questions_path, answers_path, questions_dir, figures_dir)
        result = transcribe_all_mathpix_parallel(mapping, crop_by_qnum, model)

        output_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'mathpix_output.xlsx')
        write_questions_excel(result, output_excel)
        return send_file(output_excel, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='mathpix_output.xlsx')

    except RuntimeError as e:
        return jsonify({"error": str(e)}), 501
    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        for path in (questions_path, answers_path):
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@app.route('/api/validate', methods=['POST'])
def validate_qa():
    """Validate an Excel Q&A sheet against questions_pdf and answers_pdf."""
    questions_path = answers_path = excel_path = None
    upload_folder = app.config['UPLOAD_FOLDER']
    try:
        for field in ('questions_pdf', 'answers_pdf', 'excel'):
            if field not in request.files:
                return jsonify({"error": f"Missing required field: '{field}'"}), 400

        questions_file = request.files['questions_pdf']
        answers_file   = request.files['answers_pdf']
        excel_file     = request.files['excel']

        if not (allowed_file(questions_file.filename) and allowed_file(answers_file.filename)):
            return jsonify({"error": "questions_pdf and answers_pdf must be PDF files"}), 400
        if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "excel must be an .xlsx or .xls file"}), 400

        questions_path = os.path.join(upload_folder, secure_filename(questions_file.filename))
        answers_path   = os.path.join(upload_folder, secure_filename(answers_file.filename))
        excel_path     = os.path.join(upload_folder, secure_filename(excel_file.filename))
        questions_file.save(questions_path)
        answers_file.save(answers_path)
        excel_file.save(excel_path)

        base_dir         = os.getcwd()
        questions_dir    = os.path.join(base_dir, 'questions')
        check_images_dir = os.path.join(base_dir, 'check_images')
        os.makedirs(questions_dir, exist_ok=True)
        os.makedirs(check_images_dir, exist_ok=True)

        crop_by_qnum     = crop_questions_from_pdf(questions_path, questions_dir)
        q_figures_by_num = extract_figures_per_question(questions_path, check_images_dir)

        if not crop_by_qnum:
            return jsonify({"error": "No questions could be detected in questions_pdf"}), 422

        processor    = PDFProcessor(questions_path, answers_path)
        answers_list = processor.parse_answers(processor.extract_text_from_pdf(answers_path))

        df   = pd.read_excel(excel_path)
        norm = normalise_cols(df)
        q_num_col  = pick_col(norm, ['question_number', 'question_num', 'q_num', 'qnum', 'num'])
        q_text_col = pick_col(norm, ['question_text', 'question', 'q_text', 'qtext'])
        ans_col    = pick_col(norm, ['answer', 'correct_answer', 'answers'])
        fig_col    = pick_col(norm, ['figures', 'figure', 'figure_names'])

        if not q_num_col or not q_text_col:
            return jsonify({"error": "Excel must have question_number and question_text columns"}), 422

        excel_by_qnum: dict = {}
        for _, row in df.iterrows():
            raw_num = row.get(q_num_col)
            if raw_num is None:
                continue
            try:
                q_num = int(raw_num)
            except (ValueError, TypeError):
                continue
            excel_by_qnum[q_num] = {
                "question_text": "" if pd.isna(row[q_text_col]) else str(row[q_text_col]),
                "answer":        ("" if pd.isna(row[ans_col])   else str(row[ans_col]))   if ans_col else "",
                "figures":       ("" if pd.isna(row[fig_col])   else str(row[fig_col]))   if fig_col else "",
            }

        results = []
        for q_num in sorted(excel_by_qnum.keys()):
            exc_entry  = excel_by_qnum[q_num]
            excel_q    = exc_entry["question_text"]
            excel_a    = exc_entry["answer"]
            excel_figs = exc_entry["figures"]
            pdf_a          = sanitize(latex_to_unicode(
                answers_list[q_num - 1] if (q_num - 1) < len(answers_list) else "N/A"
            ))
            crop_path      = crop_by_qnum.get(q_num)
            extracted_figs = q_figures_by_num.get(q_num, [])

            if crop_path and os.path.exists(crop_path):
                vlm_result      = vlm_compare_question(crop_path, excel_q)
                match_type      = "VLM"
                match_score     = round(vlm_result.get("confidence", 0.0) * 100)
                q_match         = None if vlm_result.get("error") else vlm_result.get("match", False)
                issues          = vlm_result.get("issues") or []
                reason          = "; ".join(issues) if issues else ""
                image_fig_count = 0 if vlm_result.get("error") else int(vlm_result.get("figure_count", 0))
            else:
                match_type = "No Image"; match_score = 0; q_match = None
                reason = "No question image available"; image_fig_count = 0

            excel_fig_names = [n.strip() for n in excel_figs.split(",") if n.strip()]
            figures_match   = len(extracted_figs) == len(excel_fig_names)
            if not figures_match:
                fig_reason = (f"Figure mismatch: image has {len(extracted_figs)} figure(s), "
                              f"Excel lists {len(excel_fig_names)}")
                reason = f"{reason}; {fig_reason}" if reason else fig_reason

            ans_sim = _fuzz.ratio(pdf_a.strip(), excel_a.strip())
            if q_match is None:     status = "Manual Review"
            elif not q_match:       status = "Incorrect"
            elif not excel_a:       status = "Manual Review"
            elif ans_sim >= 80:     status = "Correct"
            else:                   status = "Incorrect"

            results.append({
                "q_num":             q_num,
                "excel_question":    excel_q,
                "pdf_answer":        pdf_a,
                "excel_answer":      excel_a,
                "match_type":        match_type,
                "match_score":       match_score,
                "status":            status,
                "reason":            reason,
                "image_fig_count":   image_fig_count,
                "validated_figures": ", ".join(os.path.basename(p) for p in extracted_figs),
                "excel_figures":     excel_figs,
                "figures_match":     figures_match,
            })

        if not results:
            return jsonify({"error": "No matching questions found between Excel and PDF"}), 422

        output_excel = os.path.join(upload_folder, 'validation_output.xlsx')
        build_validation_excel(results, output_excel)
        return send_file(output_excel, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'validation_{len(results)}q.xlsx')

    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        for path in (questions_path, answers_path, excel_path):
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@app.route('/api/general-purpose-extraction', methods=['POST'])
def general_purpose_extraction():
    pdf_path = tmp_dir = None
    try:
        if 'pdf' not in request.files:
            return jsonify({"error": "Missing required file: pdf"}), 400
        pdf_file = request.files['pdf']
        if not pdf_file.filename or not pdf_file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Uploaded file must be a PDF"}), 400

        model    = request.form.get("model", "qwen2.5vl:7b")
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(pdf_file.filename))
        pdf_file.save(pdf_path)
        tmp_dir = tempfile.mkdtemp()

        # ── 1. Classify pages and detect layout ────────────────────────────────
        page_images = pdf_pages_to_png(pdf_path, tmp_dir, prefix="page")
        pages       = classify_pages(pdf_path, page_images)

        q_indices    = [p["page"] - 1 for p in pages if p["page_type"] == "questions"]
        s_indices    = [p["page"] - 1 for p in pages if p["page_type"] == "solutions"]
        misc_indices = [p["page"] - 1 for p in pages if p["page_type"] == "misc"]
        q_layout     = {p["page"] - 1: p["layout_type"] for p in pages if p["page_type"] == "questions"}
        s_layout     = {p["page"] - 1: p["layout_type"] for p in pages if p["page_type"] == "solutions"}

        q_images_dir = os.path.join(os.getcwd(), "question_images")
        s_images_dir = os.path.join(os.getcwd(), "solution_images")
        os.makedirs(q_images_dir, exist_ok=True)
        os.makedirs(s_images_dir, exist_ok=True)

        # ── 2. Crop individual questions, transcribe with VLM ──────────────────
        q_crops = crop_questions_from_pages(
            pdf_path, q_indices, q_images_dir,
            prefix="question", layout_by_page=q_layout,
        )
        q_texts = transcribe_crops(q_crops, QUESTION_TRANSCRIBE_PROMPT, model)

        # ── 3. Crop individual solutions, transcribe with VLM ──────────────────
        s_crops = crop_solutions_from_pages(
            pdf_path, s_indices, s_images_dir,
            prefix="solution", layout_by_page=s_layout,
        )
        s_texts = transcribe_crops(s_crops, SOLUTION_CROP_PROMPT, model)

        # ── 4. Extract compact answer key (e.g. "1.(a)  2.(c)  3.(b)") ────────
        key_answers, _ = extract_answer_keys_from_regions(pdf_path, s_indices)

        # ── 5. Map embedded figures to question numbers ────────────────────────
        figs_dir  = os.path.join(tmp_dir, "figures")
        q_fig_map = map_figures_to_questions_on_pages(
            pdf_path, q_indices,
            extract_figures_from_pages(pdf_path, q_indices, figs_dir),
        )

        # ── 6. Build Excel with separate answer_key and solution columns ───────
        # Only include rows that have at least some data; number them 1, 2, 3…
        # so gaps in image-file numbering don't appear as jumps in the sheet.
        all_q_nums = sorted(set(q_texts) | set(key_answers) | set(s_texts))
        excel_rows = []
        for row_idx, q_num in enumerate(all_q_nums, start=1):
            qt = sanitize(latex_to_unicode(q_texts.get(q_num, "")))
            ak = sanitize(latex_to_unicode(key_answers.get(q_num, "")))
            sol = sanitize(latex_to_unicode(s_texts.get(q_num, "")))
            figs = ", ".join(os.path.basename(p) for p in q_fig_map.get(q_num, []))
            if not any([qt, ak, sol, figs]):
                continue
            excel_rows.append({
                "question_num":  str(row_idx),
                "question_text": qt,
                "figures":       figs,
                "answer_key":    ak,
                "solution":      sol,
            })

        output_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'general_extraction_output.xlsx')
        write_general_extraction_excel(excel_rows, output_excel)

        return send_file(
            output_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='general_extraction_output.xlsx',
        )

    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        if pdf_path and os.path.exists(pdf_path):
            try: os.remove(pdf_path)
            except Exception: pass
        if tmp_dir and os.path.exists(tmp_dir):
            try: shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception: pass


@app.route('/api/arihant-pdfs', methods=['POST'])
def arihant_pdfs():
    pdf_path = None
    try:
        if 'pdf' not in request.files:
            return jsonify({"error": "Missing required file: pdf"}), 400
        pdf_file = request.files['pdf']
        if not pdf_file.filename or not pdf_file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Uploaded file must be a PDF"}), 400

        model_name = request.form.get("model", "claude-haiku")
        pdf_path   = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(pdf_file.filename))
        pdf_file.save(pdf_path)

        config   = PipelineConfig(vlm_model=model_name, dpi=300)
        pipeline = GeneralPurposeExtractionPipeline(config)
        result   = pipeline.run(pdf_path)

        exporter     = Exporter(config)
        export_paths = exporter.export(result, output_dir=app.config['UPLOAD_FOLDER'])
        output_excel = export_paths["excel_path"]

        return send_file(
            output_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='arihant_extraction.xlsx',
        )

    except Exception as e:
        return jsonify({"error": f"Processing error: {str(e)}"}), 500
    finally:
        if pdf_path and os.path.exists(pdf_path):
            try: os.remove(pdf_path)
            except Exception: pass


if __name__ == '__main__':
    app.run(debug=True, host='localhost', port=5000)
